from __future__ import annotations

import csv
import json
import sqlite3
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from .db import DatabaseController
from .validation import quote_identifier, validate_identifier


def export_table_csv(db: DatabaseController, table_name: str, path: str | Path) -> Path:
    return _export_table_delimited(db, table_name, path, ",")


def export_table_tsv(db: DatabaseController, table_name: str, path: str | Path) -> Path:
    return _export_table_delimited(db, table_name, path, "\t")


def export_table_json(db: DatabaseController, table_name: str, path: str | Path) -> Path:
    columns, rows = db.fetch_table_data(table_name)
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    data = [
        OrderedDict((column, row.get(column)) for column in columns)
        for row in rows
    ]
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path


def export_table_xlsx(db: DatabaseController, table_name: str, path: str | Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _worksheet_title(table_name, set())
    _write_table_sheet(db, table_name, worksheet)
    return _save_workbook(workbook, path)


def export_all_xlsx(db: DatabaseController, path: str | Path) -> Path:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    used_titles: set[str] = set()
    for table_name in db.list_tables():
        title = _worksheet_title(table_name, used_titles)
        used_titles.add(title)
        worksheet = workbook.create_sheet(title=title)
        _write_table_sheet(db, table_name, worksheet)

    if not workbook.worksheets:
        workbook.create_sheet(title="Tables")
    return _save_workbook(workbook, path)


def export_table_sql(db: DatabaseController, table_name: str, path: str | Path) -> Path:
    return export_all_sql(db, path, [table_name])


def export_all_sql(db: DatabaseController, path: str | Path, table_names: list[str] | None = None) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    selected_tables = table_names if table_names is not None else db.list_tables()

    lines = [
        "-- SQL Data Controller ordered export",
        "BEGIN TRANSACTION;",
    ]
    for table_name in selected_tables:
        validate_identifier(table_name, "table name")
        columns, rows = db.fetch_table_data(table_name)
        lines.append(_create_table_sql(table_name, columns))
        for row in rows:
            lines.append(_insert_row_sql(table_name, columns, row))
    lines.append("COMMIT;")
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return output_path


def backup_sqlite_database(db: DatabaseController, path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    db.connection.commit()
    temp_path = output_path.with_name(f".{output_path.name}.backup_tmp")
    _remove_sqlite_backup_files(temp_path)
    try:
        with sqlite3.connect(temp_path) as destination:
            db.connection.backup(destination)
        _finalize_sqlite_backup_file(temp_path)
        _remove_sqlite_backup_files(output_path)
        temp_path.replace(output_path)
    except sqlite3.Error:
        _remove_sqlite_backup_files(temp_path)
        output_path.write_bytes(db.connection.serialize())
        _finalize_sqlite_backup_file(output_path)
    return output_path


def _export_table_delimited(
    db: DatabaseController,
    table_name: str,
    path: str | Path,
    delimiter: str,
) -> Path:
    columns, rows = db.fetch_table_data(table_name)
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, delimiter=delimiter)
        writer.writerow(columns)
        for row in rows:
            writer.writerow(["" if row.get(column) is None else row.get(column) for column in columns])
    return output_path


def _write_table_sheet(db: DatabaseController, table_name: str, worksheet) -> None:
    columns, rows = db.fetch_table_data(table_name)
    worksheet.append(columns)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    for row in rows:
        worksheet.append([row.get(column) for column in columns])


def _save_workbook(workbook: Workbook, path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _create_table_sql(table_name: str, columns: list[str]) -> str:
    definitions = ",\n    ".join(_column_definition(column) for column in columns)
    return f"CREATE TABLE {quote_identifier(table_name)} (\n    {definitions}\n);"


def _column_definition(column: str) -> str:
    if column == "id":
        return f"{quote_identifier(column)} INTEGER PRIMARY KEY AUTOINCREMENT"
    if column == "_row_name":
        return f"{quote_identifier(column)} TEXT UNIQUE"
    return f"{quote_identifier(column)} TEXT"


def _insert_row_sql(table_name: str, columns: list[str], row: dict[str, Any]) -> str:
    column_sql = ", ".join(quote_identifier(column) for column in columns)
    values_sql = ", ".join(_sql_literal(row.get(column)) for column in columns)
    return f"INSERT INTO {quote_identifier(table_name)} ({column_sql}) VALUES ({values_sql});"


def _sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int | float):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def _worksheet_title(table_name: str, used_titles: set[str]) -> str:
    base = table_name[:31] or "Table"
    title = base
    index = 2
    while title in used_titles:
        suffix = f"_{index}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    return title


def _finalize_sqlite_backup_file(path: Path) -> None:
    with sqlite3.connect(path) as connection:
        connection.execute("PRAGMA journal_mode = TRUNCATE")


def _remove_sqlite_backup_files(path: Path) -> None:
    for candidate in (
        path,
        path.with_name(f"{path.name}-journal"),
        path.with_name(f"{path.name}-wal"),
        path.with_name(f"{path.name}-shm"),
    ):
        try:
            candidate.unlink()
        except (FileNotFoundError, PermissionError):
            pass
