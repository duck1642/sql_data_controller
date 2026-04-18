from __future__ import annotations

import csv
import json
import shutil
import sqlite3
import unittest
import uuid
from pathlib import Path

from openpyxl import load_workbook

from app.db import DatabaseController
from app.exporters import (
    backup_sqlite_database,
    export_all_sql,
    export_all_xlsx,
    export_table_csv,
    export_table_json,
    export_table_sql,
    export_table_tsv,
    export_table_xlsx,
)


class ExporterTests(unittest.TestCase):
    def setUp(self) -> None:
        temp_root = Path.cwd() / ".test_tmp"
        temp_root.mkdir(exist_ok=True)
        self.temp_dir = temp_root / uuid.uuid4().hex
        self.temp_dir.mkdir()
        self.db = DatabaseController(self.temp_dir / "database.sqlite")

    def tearDown(self) -> None:
        self.db.close()
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def create_ordered_customer_table(self) -> tuple[int, int]:
        self.db.create_table("customers")
        self.db.add_column("customers", "name")
        self.db.add_column("customers", "email")
        first_id = self.db.add_row("customers", "first")
        second_id = self.db.add_row("customers", "second")
        self.db.update_cell("customers", first_id, "name", "Ali")
        self.db.update_cell("customers", first_id, "email", "ali@example.com")
        self.db.update_cell("customers", second_id, "name", "Ayse")
        self.db.reorder_rows("customers", [second_id, first_id])
        self.db.reorder_columns("customers", ["name", "email", "_row_name", "id"])
        return first_id, second_id

    def test_csv_and_tsv_use_reordered_rows_and_columns(self) -> None:
        self.create_ordered_customer_table()

        csv_path = export_table_csv(self.db, "customers", self.temp_dir / "customers.csv")
        tsv_path = export_table_tsv(self.db, "customers", self.temp_dir / "customers.tsv")

        with csv_path.open(newline="", encoding="utf-8") as handle:
            csv_rows = list(csv.reader(handle))
        with tsv_path.open(newline="", encoding="utf-8") as handle:
            tsv_rows = list(csv.reader(handle, delimiter="\t"))

        expected = [
            ["name", "email", "_row_name", "id"],
            ["Ayse", "", "second", "2"],
            ["Ali", "ali@example.com", "first", "1"],
        ]
        self.assertEqual(csv_rows, expected)
        self.assertEqual(tsv_rows, expected)

    def test_json_preserves_ordered_keys_and_null_values(self) -> None:
        self.create_ordered_customer_table()

        json_path = export_table_json(self.db, "customers", self.temp_dir / "customers.json")
        data = json.loads(json_path.read_text(encoding="utf-8"))

        self.assertEqual(list(data[0].keys()), ["name", "email", "_row_name", "id"])
        self.assertEqual(data[0]["name"], "Ayse")
        self.assertIsNone(data[0]["email"])

    def test_xlsx_current_table_export_writes_headers_and_rows(self) -> None:
        self.create_ordered_customer_table()

        xlsx_path = export_table_xlsx(self.db, "customers", self.temp_dir / "customers.xlsx")
        workbook = load_workbook(xlsx_path)
        worksheet = workbook["customers"]

        self.assertEqual([cell.value for cell in worksheet[1]], ["name", "email", "_row_name", "id"])
        self.assertEqual([cell.value for cell in worksheet[2]], ["Ayse", None, "second", 2])
        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertTrue(worksheet["A1"].font.bold)

    def test_xlsx_all_tables_export_creates_sheet_per_user_table(self) -> None:
        self.create_ordered_customer_table()
        self.db.create_table("orders")

        xlsx_path = export_all_xlsx(self.db, self.temp_dir / "all.xlsx")
        workbook = load_workbook(xlsx_path)

        self.assertEqual(set(workbook.sheetnames), {"customers", "orders"})

    def test_clean_sql_export_excludes_metadata_and_uses_order(self) -> None:
        self.create_ordered_customer_table()

        sql_path = export_all_sql(self.db, self.temp_dir / "ordered.sql")
        text = sql_path.read_text(encoding="utf-8")

        self.assertNotIn("_sdc_", text)
        create_sql = text[text.index("CREATE TABLE") : text.index(");") + 2]
        self.assertLess(create_sql.index('"name" TEXT'), create_sql.index('"_row_name" TEXT UNIQUE'))
        self.assertLess(create_sql.index('"_row_name" TEXT UNIQUE'), create_sql.index('"id" INTEGER PRIMARY KEY AUTOINCREMENT'))
        self.assertLess(text.index("'Ayse'"), text.index("'Ali'"))

    def test_current_table_sql_export_is_executable(self) -> None:
        self.create_ordered_customer_table()

        sql_path = export_table_sql(self.db, "customers", self.temp_dir / "customers.sql")
        text = sql_path.read_text(encoding="utf-8")

        with sqlite3.connect(":memory:") as connection:
            connection.executescript(text)
            count = connection.execute('SELECT COUNT(*) FROM "customers"').fetchone()[0]

        self.assertEqual(count, 2)

    def test_sqlite_backup_creates_readable_internal_copy(self) -> None:
        self.create_ordered_customer_table()

        backup_path = backup_sqlite_database(self.db, self.temp_dir / "backup.sqlite")

        with sqlite3.connect(backup_path) as connection:
            table_names = {
                row[0]
                for row in connection.execute("SELECT name FROM sqlite_master WHERE type = 'table'")
            }

        self.assertIn("customers", table_names)
        self.assertIn("_sdc_row_order", table_names)


if __name__ == "__main__":
    unittest.main()
